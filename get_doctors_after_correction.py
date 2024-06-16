import random
from doctor_class_correction import doctor
import openpyxl
import sqlite3

def get_mas_of_doc():
    all_mas = []
    workbook = openpyxl.load_workbook('graph_file.xlsx')
    worksheet = workbook.active
    all_new = []

    for i in range(1, worksheet.max_row):
        a = 0
        all_mas.append([])
        all_new.append([[], [], []])
        for col in worksheet.iter_cols(0, 39):
            if (col[i].value == None) & (a == 0):
                break

            if a == 0:
                all_mas[i-1].append(col[i].value)
            if a == 1:
                if col[i].value == 'КТ':
                    all_mas[i - 1].append(['КТ'])
                    all_mas[i - 1][1].append('КТ1')
                    all_mas[i - 1][1].append('КТ2')

                elif col[i].value == 'МРТ':
                    all_mas[i - 1].append(['МРТ'])
                    all_mas[i - 1][1].append('МРТ1')
                    all_mas[i - 1][1].append('МРТ2')

                else:
                    all_mas[i - 1].append([col[i].value])

            if (a == 2) & (col[i].value != None) & (col[i].value != '-'):
                tx = col[i].value.split(',')
                for mes in tx:
                    if mes.strip() == 'Денс':
                        all_mas[i - 1][1].append('Денситометрия')
                    elif mes.strip() == 'Денситометрия':
                        all_mas[i - 1][1].append('Денситометрия')
                    elif mes.strip() == 'РГ':
                        all_mas[i - 1][1].append('РГ')
                    elif mes.strip() == 'КТ':
                        all_mas[i - 1][1].append('КТ')
                        all_mas[i - 1][1].append('КТ1')
                        all_mas[i - 1][1].append('КТ2')
                    elif mes.strip() == 'ММГ':
                        all_mas[i - 1][1].append('ММГ')
                    elif mes.strip() == 'ФЛГ':
                        all_mas[i - 1][1].append('ФЛГ')
                    elif mes.strip() == 'МРТ':

                        all_mas[i - 1][1].append('МРТ')
                        all_mas[i - 1][1].append('МРТ1')
                        all_mas[i - 1][1].append('МРТ2')





            if a == 3:
                all_mas[i-1].append(col[i].value)


            if (a == 6) | (a == 7) | (a == 8)| (a == 9)| (a == 10) | (a == 11)| (a == 12)| (a == 13)    | (a == 14)| (a == 15)| (a == 16)| (a == 17)| (a == 18)|(a == 19)|(a == 20)|(a == 22)| (a == 23)|(a == 24)|(a == 25)| (a == 26)|(a == 27)|(a == 28)| (a == 29)|(a == 30)| (a == 31)|(a == 32)|(a == 33)|(a == 34)|(a == 35)|(a == 36)|(a == 37)|(a == 38):
                if col[i].value == None or col[i].value == '' or col[i].value == 0:
                    all_new[i-1][0].append(0)
                    all_new[i-1][1].append(0)
                    all_new[i-1][2].append(0)
                else:
                    all_new[i - 1][0].append(1)
                    all_new[i - 1][1].append(col[i+3].value)
                    all_new[i - 1][2].append(str(col[i].value))



            a+=1
        #print(all_new[i-1])
    for_return = [x for x in all_mas if x != []]
    for_new = [x for x in all_new if x != [[],[],[]]]
    print(for_new)
    return for_return, for_new

def create_doctors():
    all_doctors, all_new = get_mas_of_doc()
    mas_of_two_param = import_weeks_and_sched()
    all_docs = []
    i = 0
    for doc in all_doctors:
        #print(str(doc[0]) + ' '+ str(doc[1]) + ' '+   str(doc[2]) + ' '+   str(mas_of_two_param[i][0]) + ' '+   str(all_new[i][0]) + ' '+   str(mas_of_two_param[i][1]) + ' '+   str(all_new[i][1]) + ' '+   str(all_new[i][2]))
        all_docs.append(doctor(doc[0], doc[1], doc[2], mas_of_two_param[i][0], all_new[i][0], mas_of_two_param[i][1], all_new[i][1], all_new[i][2],True))                #7
        i+=1
    for doc in all_docs:
        doc.vivod()
    return all_docs


def import_weeks_and_sched():
    mas_of_two_param = []
    sqlite_connection = sqlite3.connect('db.sqlite3')
    cursor = sqlite_connection.cursor()
    cursor.execute(f"SELECT COUNT(*) FROM accounts_employee")

    # Получение результата
    row_count = cursor.fetchone()[0]
    sqlite_select_query = """SELECT * from accounts_employee"""
    cursor.execute(sqlite_select_query)
    records = cursor.fetchmany(row_count)
    j = 0
    for i in range(0, len(records)):
        if records[i][1] ==1:

            b = [records[i][5][1], records[i][5][4]]
            A = records[i][6][2:-2]

            # Разделяем строку по '], ['
            list_str = A.split('], [')

            # Преобразуем каждую строку внутри вложенного списка в список целых чисел
            array_2d = [[int(num) for num in sublist.split(', ')] for sublist in list_str]


            mas_of_two_param.append([b,array_2d])             #5,6
    cursor.close()

    sqlite_connection.close()
    return mas_of_two_param


create_doctors()






