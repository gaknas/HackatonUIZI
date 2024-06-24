import sys
import os
import pandas as pd
from statsmodels.tsa.arima.model import ARIMA
import openpyxl
import shutil
import sqlite3
from pmdarima import auto_arima
from datetime import datetime
start_time = datetime.now()

# Загрузка Excel файла
source_file = sys.argv[1]
df = pd.read_excel(source_file)
# Замена всех пустых ячеек на 0
df.fillna(0, inplace=True)
# Сохранение изменений в Excel файле
df.to_excel(source_file, index=False)
#source_file = 'output_file_count.xlsx'
destination_file = 'output_file_count.xlsx'
shutil.copy2(source_file, destination_file)

# Загрузка данных из Excel-файла
df = pd.read_excel(source_file)
# Удаление строк, в которых значение во второй ячейке равно 1
df = df[df.iloc[:, 1] != 1]
# Сохранение измененного файла
df.to_excel('Количество исследований по неделям без 1 .xlsx', index=False)


t = 0
array_ov = []
for iiii in range(0, 10):
    # Чтение Excel файла
    df = pd.read_excel('Количество исследований по неделям без 1 .xlsx')
    # Выбор третьего столбца без первых двух строк
    column_to_copy = df.iloc[2:, iiii+2]
    # Запись данных в txt файл
    with open(str(iiii+3) + '_Количество исследований по неделям без 1 .txt', 'w') as file:
        for item in column_to_copy:
            file.write(str(item) + '\n')

    # Чтение данных из txt файла
    data = pd.read_csv(str(iiii+3) + '_Количество исследований по неделям без 1 .txt', header=None)
    # Преобразование данных в одномерный массив
    data = data.values.flatten()
    # Подбор оптимальных параметров ARIMA
    model = auto_arima(data, seasonal=False, trace=True)
    # Печать оптимальных параметров ARIMA
    print(model.summary())
    p = model.get_params()['order'][0]
    d = model.get_params()['order'][1]
    q = model.get_params()['order'][2]


    # Загружаем данные из файла
    data = pd.read_csv(str(iiii+3) + '_Количество исследований по неделям без 1 .txt', header=None, names=['value'])
    # Обучаем модель ARIMA
    model = ARIMA(data, order=(p, d, q))
    model_fit = model.fit()
    # Предсказываем значения временного ряда
    predictions = model_fit.forecast(steps=5)
    # Выводим предсказанные значения
    arr = []
    # Запись данных в txt файл
    with open(str(iiii+3) + '_Предсказание.txt', 'w') as file:
        for item in predictions:
            file.write(str(item) + '\n')
            arr.append(int(item))
            array_ov.append(int(item))


    file_path = 'output_file_count.xlsx'
    data = pd.read_excel(file_path)
    num_rows = data.shape[0]
    if iiii == 0:
        num_rows1 = num_rows
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    for i in range(0, len(arr)):
        sheet.cell(row=num_rows1 + 2 + i, column=iiii+3, value=arr[i])
    wb.save(file_path)


rows = []
df = pd.read_excel('output_file_count.xlsx')
# Поиск строк, в которых значение во втором столбце равно 1
rows_with_1 = df[df.iloc[:, 1] == 1].index
for row in rows_with_1:
    rows.append(row)


# Открываем Excel файл
wb = openpyxl.load_workbook('output_file_count.xlsx')
sheet = wb.active
# Находим последнюю заполненную строку
last_row = sheet.max_row
for i in range(last_row, 0, -1):
    if sheet.cell(row=i, column=2).value is not None:
        last_value = sheet.cell(row=i, column=2).value
        a_2024 = (sheet.cell(row=i, column=1).value)
        break



file_path = 'output_file_count.xlsx'
data = pd.read_excel(file_path)
num_rows = data.shape[0]
wb = openpyxl.load_workbook(file_path)
sheet = wb.active
last_value += 1
for i in range(0, len(arr)):
    if last_value == 53:
        last_value = 1
        sheet.cell(row=num_rows + 2 + i - 5, column=2, value=1)
        a_2024 += 1
    if last_value == 1:
        for j in range(0, 10):
            sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=df.iloc[rows[len(rows)-1], j+2])
    if last_value == 3:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.06))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.06))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.07))))
    if last_value == 4:
        for j in range(0, 10):
            if j == 2:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.13))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.25))))
    if last_value == 5:
        for j in range(0, 10):
            if j == 2:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.056))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.23))))
    if last_value == 6:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]+(array_ov[i+j*5]*0.16))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.09))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.06))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.20))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.1))))
    if last_value == 7:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.08))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.12))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.06))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.09))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*1.25))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.12))))
    if last_value == 8:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.22))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.25))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.18))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.13))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.16))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.34))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.2))))
    if last_value == 9:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]+(array_ov[i+j*5]*0.71))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.31))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.64))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.35))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.54))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.20))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.34))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.43))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.56))))
    if last_value == 10:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.28))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]+(array_ov[i+j*5]*0.29))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.23))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.27))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.27))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.19))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.16))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.28))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.24))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.28))))
    if last_value == 11:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]+(array_ov[i+j*5]*0.51))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.17))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.48))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.49))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.43))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.27))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.25))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.33))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.49))))
    if last_value == 12:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]+(array_ov[i+j*5]*0.23))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.05))))
    if last_value == 14:
        for j in range(0, 10):
            if j == 3:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.067))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]))))
    if last_value == 15:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]+(array_ov[i+j*5]*0.02))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.45))))
    if last_value == 16:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.05))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.09))))
    if last_value == 17:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]-(array_ov[i+j*5]*0.06))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=int(abs(array_ov[i+j*5]+(array_ov[i+j*5]*0.03))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.10))))
    if last_value == 18:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.16))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.07))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.27))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.29))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.26))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.16))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.22))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.22))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.29))))
    if last_value == 19:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.20))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.07))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.15))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.83))))
    if last_value == 20:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.31))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.67))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.47))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.55))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.30))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.43))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.34))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.45))))
    if last_value == 21:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.31))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.06))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.28))))
    if last_value == 22:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.31))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.06))))
    if last_value == 23:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.03))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.059))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.077))))
    if last_value == 24:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.09))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.03))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.13))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.15))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.16))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.14))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.275))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.13))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.19))))
    if last_value == 25:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.10))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.12))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.18))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.27))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.11))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.087))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.16))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.12))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.18))))
    if last_value == 26:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.15))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.05))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.14))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.25))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.096))))
    if last_value == 27:
        for j in range(0, 10):
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.875))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.1))))
    if last_value == 28:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.079))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.15))))
    if last_value == 29:
        for j in range(0, 10):
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.11))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.07))))
    if last_value == 30:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.06))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.06))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.38))))
    if last_value == 31:
        for j in range(0, 10):
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.07))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*1.875))))
    if last_value == 32:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.11))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.087))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.09))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.42))))
    if last_value == 33:
        for j in range(0, 10):
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.064))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.06))))
    if last_value == 34:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.07))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.04))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.087))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.14))))
    if last_value == 35:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.04))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.05))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.09))))
    if last_value == 36:
        for j in range(0, 10):
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.09))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.34))))
    if last_value == 37:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.08))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.26))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.06))))
    if last_value == 38:
        for j in range(0, 10):
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.045))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.05))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.05))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.89))))
    if last_value == 39:
        for j in range(0, 10):
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.063))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.16))))
    if last_value == 40:
        for j in range(0, 10):
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.116))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.063))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.16))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.3))))
    if last_value == 41:
        for j in range(0, 10):
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.70))))
    if last_value == 42:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.05))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.07))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.07))))
    if last_value == 43:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.04))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.093))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.086))))
    if last_value == 44:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.11))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.08))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.09))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.09))))
    if last_value == 45:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.16))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.06))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.085))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.22))))

    if last_value == 46:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.10))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.12))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.10))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.07))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.12))))

    if last_value == 47:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.5))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.06))))
    if last_value == 48:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.06))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.06))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.1))))

    if last_value == 49:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] + (array_ov[i + j * 5] * 0.03))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.06))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.61))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.08))))
    if last_value == 50:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.03))))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] + (array_ov[i+j*5]*0.09))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.065))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.1))))
    if last_value == 51:
        for j in range(0, 10):
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.199))))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.24))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.10))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.16))))
    if last_value == 52:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.19))))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3,
                           value=int(abs(array_ov[i + j * 5] - (array_ov[i + j * 5] * 0.1))))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.15))))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.26))))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.15))))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.61))))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.0))))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=int(abs(array_ov[i + j * 5] - (array_ov[i+j*5]*0.25))))

    else:
        sheet.cell(row=num_rows + 2 + i-5, column=2, value=abs(last_value))
    sheet.cell(row=num_rows + 2 + i - 5, column=1, value=a_2024)
    last_value += 1
wb.save(file_path)


"""


file_path = 'output_file_count.xlsx'
data = pd.read_excel(file_path)
num_rows = data.shape[0]
wb = openpyxl.load_workbook(file_path)
sheet = wb.active
last_value += 1
for i in range(0, len(arr)):
    if last_value == 53:
        last_value = 1
        sheet.cell(row=num_rows + 2 + i - 5, column=2, value=1)
        a_2024 += 1
    if last_value == 1:
        for j in range(0, 10):
            sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=df.iloc[rows[len(rows)-1], j+2])
    if last_value == 2:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-350))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-550))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-50))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-100))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-3000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-500))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-100))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-3))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-10000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=abs(array_ov[i+j*5]-5000))
    if last_value == 8:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 450))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 1550))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 250))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 100))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 6000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 400))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 300))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 1))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 15000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 9000))
    if last_value == 18:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 150))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 550))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 110))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 100))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 2000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 100))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 10000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 5000))
    if last_value == 19:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 550))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 950))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 210))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 300))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 5000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 400))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 300))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 15000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 6000))
    if last_value == 24:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 150))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 350))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 110))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 6000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
    if last_value == 27:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 350))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 650))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 110))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 6000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
    if last_value == 28:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 350))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 650))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 110))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 6000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
    if last_value == 29:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 350))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 650))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 110))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 6000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
    if last_value == 30:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 350))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 650))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 110))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 6000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
    if last_value == 38:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 350))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] + 450))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] + 110))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] + 100))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] + 5000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5]))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 12000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
    if last_value == 52:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 350))
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 650))
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 200))
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 300))
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 100))
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 4))
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 8000))
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=abs(array_ov[i + j * 5] - 3000))
    else:
        sheet.cell(row=num_rows + 2 + i-5, column=2, value=abs(last_value))
    sheet.cell(row=num_rows + 2 + i - 5, column=1, value=a_2024)
    last_value += 1
wb.save(file_path)
"""


os.remove("3_Количество исследований по неделям без 1 .txt")
os.remove("4_Количество исследований по неделям без 1 .txt")
os.remove("5_Количество исследований по неделям без 1 .txt")
os.remove("6_Количество исследований по неделям без 1 .txt")
os.remove("7_Количество исследований по неделям без 1 .txt")
os.remove("8_Количество исследований по неделям без 1 .txt")
os.remove("9_Количество исследований по неделям без 1 .txt")
os.remove("10_Количество исследований по неделям без 1 .txt")
os.remove("11_Количество исследований по неделям без 1 .txt")
os.remove("12_Количество исследований по неделям без 1 .txt")
os.remove("3_Предсказание.txt")
os.remove("4_Предсказание.txt")
os.remove("5_Предсказание.txt")
os.remove("6_Предсказание.txt")
os.remove("7_Предсказание.txt")
os.remove("8_Предсказание.txt")
os.remove("9_Предсказание.txt")
os.remove("10_Предсказание.txt")
os.remove("11_Предсказание.txt")
os.remove("12_Предсказание.txt")


# Чтение данных из Excel файла
df = pd.read_excel('output_file_count.xlsx')
# Подключение к базе данных SQLite
conn = sqlite3.connect('db.sqlite3')
cur = conn.cursor()
# Запись данных из Excel файла в базу данных
df.to_sql('accounts_excelmodel', conn, if_exists='replace', index=True)
# Завершение работы с базой данных
conn.commit()
conn.close()


end_time = datetime.now()
print('Duration: {}'.format(end_time - start_time))
os.remove('output_file_count.xlsx')
os.remove('Количество исследований по неделям без 1 .xlsx')
