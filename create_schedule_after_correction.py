import get_doctors_after_correction
from doctor_class_correction import doctor
import numpy as np
import openpyxl
import os
import sqlite3
import pandas as pd
import datetime



def check_for_overload(activities, schetchik, amount_for_month):
    mas = {}
    for key in schetchik.keys():
        mas[key] = amount_for_month[key] - schetchik[key] * activities[key] * 5
    return mas

def make_schetchik(all_docs, schetchik):
    new_schet = {}
    for key in schetchik.keys():
        new_schet[key] = 0

    for key in schetchik.keys():
        for doc in all_docs:
            if key in doc.get_abilities():
                new_schet[key] += 1
    return new_schet


def find_min_key(schetchik):
    min_key = min(schetchik, key=schetchik.get)
    #print(min_key)
    return min(schetchik, key=schetchik.get)

def find_doctors(all_docs, key, week, day):
    mas_of_suitable_doctors = []
    i = 0
    for doc in all_docs:
        #print(doc.get_available_days())
        #print(str(week) + '   '+ str(day) + '   ' + str(doc.get_id()))
        if doc.get_activities_per_day(week, day)>0:
            if key in doc.get_abilities() :
                mas_of_suitable_doctors.append(i)
        i+=1
    return mas_of_suitable_doctors

def find_priority_doctors(all_docs, key, week, day):
    mas_of_doctors_in_priority = [[], []]
    doctors = find_doctors(all_docs,key, week, day)
    for i in doctors:
        if len(all_docs[i].get_abilities()) == 1:
            mas_of_doctors_in_priority[0].append(i)
        else:
            mas_of_doctors_in_priority[1].append(i)

    return mas_of_doctors_in_priority

def update_schetchik(schetchik, all_docs, week,day):
    schetchik_new = {}
    for key in schetchik.keys():
        schetchik_new[key] = 0
    for doc in all_docs:
        if doc.get_activities_per_day(week, day) > 0:
            for key in doc.get_abilities():
                if key in schetchik_new.keys():
                    schetchik_new[key] += 1

    return schetchik_new

def zapolnenie_exel(all_docs,amount_for_weeks):
    workbook = openpyxl.load_workbook('doc.xlsx')
    worksheet = workbook.active
    book = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
            'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL']
    doc_id = 0
    schetchik_for_docs = 0
    days_in_month = 0
    for week in range(0, len(amount_for_weeks)):
        for day in range(0, len(all_docs[0].get_available_days()[week])):
            days_in_month += 1

    for doc in all_docs:
        date = 3
        shetchik_time = 0
        worksheet['A' + str(4*schetchik_for_docs +3)] = doc.get_id()
        schetchik_for_docs +=1

        for week in range(0, len(amount_for_weeks)):

            for day in range(0, len(all_docs[0].get_available_days()[week])):
                strin1 = str(book[date-3 +6]) + str(4 * doc_id + 3 )
                strin2 = str(book[date-3 +6]) + str(4 * doc_id + 4)
                strin3 = str(book[date-3 +6]) + str(4 * doc_id + 5)
                strin4 = str(book[date-3 +6]) + str(4 * doc_id + 6)



                if doc.get_amount_of_hours(week, day) == 12:
                    worksheet[strin1] = str('20:00')
                    worksheet[strin2] = str('9:00')
                    worksheet[strin3] = str((all_docs[doc_id].get_pereriv(week)[day] % 1 + int(all_docs[doc_id].get_pereriv(week)[day])) * 60)

                    timee = all_docs[doc_id].get_amount_of_hours(week, day) + all_docs[doc_id].get_pereriv(week)[day]

                    shetchik_time+=12
                    worksheet[strin4] = 12

                elif doc.get_amount_of_hours(week, day) == 0:
                    worksheet[strin1].value = 0
                    worksheet[strin2].value = 0
                    worksheet[strin3].value = 0
                    worksheet[strin4].value = 0
                else:
                    timee = all_docs[doc_id].get_amount_of_hours(week, day) + all_docs[doc_id].get_pereriv(week)[day]
                    worksheet[strin1] = str('8:00')
                    worksheet[strin2] = str(8 + int(timee)) + ':' + str(int((timee % 1) * 60))
                    worksheet[strin3] = str((all_docs[doc_id].get_pereriv(week)[day] % 1 + int(all_docs[doc_id].get_pereriv(week)[day])) * 60)



                    worksheet[strin4] = all_docs[doc_id].get_amount_of_hours(week, day)
                    shetchik_time += all_docs[doc_id].get_amount_of_hours(week, day)

                #print(date)

                if date-3 == 15:
                    strin = 'V' + str(4 * doc_id + 6)
                    worksheet[strin] = shetchik_time
                    shetchik_time = 0
                elif date-3 == days_in_month-1:
                    strin = 'AM' + str(4 * doc_id + 6)
                    worksheet[strin] = shetchik_time
                    shetchik_time = 0

                date = date + 1


        doc_id += 1

    workbook.save('sched_correction.xlsx')

def kostil(amount_for_week, worksheet, days_in_month, kost, book, schetchik):
    worksheet['A1'] = ''
    worksheet['B1']= ''
    worksheet['C1']= ''
    worksheet['D1']= ''
    worksheet['E1']= ''
    worksheet['F1']= ''
    worksheet['G1']= ''
    worksheet['H1']= ''
    worksheet['I1']= ''
    worksheet['J1']= ''

    strin1 = str(book[0]) + str(kost+1)
    strin2 = str(book[1]) + str(kost+1)
    strin3 = str(book[2]) + str(kost+1)
    strin4 = str(book[3]) + str(kost+1)
    strin5 = str(book[4]) + str(kost+1)
    strin6 = str(book[5]) + str(kost+1)
    strin7 = str(book[6]) + str(kost+1)
    strin8 = str(book[7]) + str(kost+1)
    strin9 = str(book[8]) + str(kost+1)
    strin10 = str(book[9]) + str(kost+1)
    keys = [key for key in schetchik.keys()]
    if int(amount_for_week[keys[0]]) > 0:
        worksheet[strin1] = str(keys[0])+":"+str(int(amount_for_week[keys[0]]))
    else:
        worksheet[strin1] = str(keys[0])+":"+str(int(0))
    if int(amount_for_week[keys[1]]) > 0:
        worksheet[strin2] = str(keys[1])+":"+str(int(amount_for_week[keys[1]]))
    else:
        worksheet[strin2] = str(keys[1])+":"+str(int(0))
    if int(amount_for_week[keys[2]]) > 0:
        worksheet[strin3] = str(keys[2])+":"+str(int(amount_for_week[keys[2]]))
    else:
        worksheet[strin3] = str(keys[2])+":"+str(int(0))
    if int(amount_for_week[keys[3]]) > 0:
        worksheet[strin4] = str(keys[3])+":"+str(int(amount_for_week[keys[3]]))
    else:
        worksheet[strin4] = str(keys[3])+":"+str(int(0))
    if int(amount_for_week[keys[4]]) > 0:
        worksheet[strin5] = str(keys[4])+":"+str(int(amount_for_week[keys[4]]))
    else:
        worksheet[strin5] = str(keys[4])+":"+str(int(0))
    if int(amount_for_week[keys[5]]) > 0:
        worksheet[strin6] = str(keys[5])+":"+str(int(amount_for_week[keys[5]]))
    else:
        worksheet[strin6] = str(keys[5])+":"+str(int(0))
    if int(amount_for_week[keys[6]]) > 0:
        worksheet[strin7] = str(keys[6])+":"+str(int(amount_for_week[keys[6]]))
    else:
        worksheet[strin7] = str(keys[6])+":"+str(int(0))
    if int(amount_for_week[keys[7]]) > 0:
        worksheet[strin8] = str(keys[7])+":"+str(int(amount_for_week[keys[7]]))
    else:
        worksheet[strin8] = str(keys[7])+":"+str(int(0))
    if int(amount_for_week[keys[8]]) > 0:
        worksheet[strin9] = str(keys[8])+":"+str(int(amount_for_week[keys[8]]))
    else:
        worksheet[strin9] = str(keys[8])+":"+str(int(0))
    if int(amount_for_week[keys[9]]) > 0:
        worksheet[strin10] = str(keys[9])+":"+str(int(amount_for_week[keys[9]]))
    else:
        worksheet[strin10] = str(keys[9])+":"+str(int(0))


def make_schedule_for_month(all_docs, activities, activities_UE, amount_for_weeks,schetchik):
    max = 300
    workbook1 = openpyxl.Workbook()
    worksheet1 = workbook1.active
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    book = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL']
    days_in_month = 0
    for week in range(0, len(amount_for_weeks)):
        for day in range(0, len(all_docs[0].get_available_days()[week])):
            days_in_month+=1
    kost = 1
    for week in range(0, len(amount_for_weeks)):

        for day in range(0, len(all_docs[0].get_available_days()[week])):
            amount_for_week = dict(amount_for_weeks[week])
            #print('start:')
            #print(amount_for_week)

            schetchik_for_day = update_schetchik(schetchik, all_docs, week,day)
            #print(schetchik_for_day)
            flag = len(schetchik_for_day)

            while flag > 0:

                min_key = find_min_key(schetchik_for_day)

                massive_of_doctors_in_priority_order = find_priority_doctors(all_docs, min_key,week,day)

                for priority_order in massive_of_doctors_in_priority_order:
                    if amount_for_week[min_key] <= 0:
                        break
                    for doc_id in priority_order:

                        if (amount_for_week[min_key] <= 0) | (all_docs[doc_id].get_activities_per_day(week, day) <= 0):
                            break
                        while all_docs[doc_id].get_activities_per_day(week, day) > 0:
                            if amount_for_week[min_key] <= 0:
                                break
                            amount_for_week[min_key] = amount_for_week[min_key] - (all_docs[doc_id].get_koef(week, day) * activities_UE[min_key])

                            #all_docs[doc_id].set_new_day_schedule(week, day, 0)                                                             #UBIRAET DEN
                            new_amount = all_docs[doc_id].get_activities_per_day(week, day) - (all_docs[doc_id].get_koef(week, day) * activities_UE[min_key])
                            #print(str(min_key) + '  '+ str(new_amount) +  '  ' + str(amount_for_week[min_key]))
                            all_docs[doc_id].set_activities_per_day(new_amount, week, day)                                                   #SKOLKO ENERGYY VRACHA OSTALOS

                        all_docs[doc_id].set_new_available_days(all_docs[doc_id].get_available_day(week, day),week, day)      #RASPISANIE DLA ZAPOLNENIA TABLIZ
                        all_docs[doc_id].set_work_per_day(min_key, week, day)                                                           #KAKOE ISSLEDOVANIE DELAET

                schetchik_for_day.pop(min_key)
                schetchik_for_day = update_schetchik(schetchik_for_day, all_docs, week,day)
                flag-=1
            #print('end:')
            #print(amount_for_week)
            kostil(amount_for_week, worksheet, days_in_month, kost, book, schetchik)
            kost += 1
            #print(schetchik_for_day)
    doc_id = 0
    schetchik_for_docs = 0
    workbook.save('kostil1.xlsx')
    df = pd.read_excel('kostil1.xlsx')
    os.remove('kostil1.xlsx')

    conn = sqlite3.connect('db.sqlite3')
    # Запись данных из Excel файла в базу данных
    df.to_sql('overload', conn, if_exists='replace', index=True)
    # Завершение работы с базой данных
    conn.commit()
    conn.close()

    worksheet1['A1'] = 'sys_user_id'
    worksheet1['B1'] = 'day_of_month'
    worksheet1['C1'] = 'time_start'
    worksheet1['D1'] = 'time_end'
    worksheet1['E1'] = 'time_break'
    worksheet1['F1'] = 'time_total'
    worksheet1['G1'] = 'research_type'
    for doc in all_docs:
        date = 1
        for week in range(0, len(amount_for_weeks)):

            for day in range(0, len(all_docs[0].get_available_days()[week])):
                strin1 = str(book[0]) + str(days_in_month * doc_id + 1 + date)
                strin2 = str(book[1]) + str(days_in_month * doc_id + 1 + date)
                strin3 = str(book[2]) + str(days_in_month * doc_id + 1 + date)
                strin4 = str(book[3]) + str(days_in_month * doc_id + 1 + date)
                strin5 = str(book[4]) + str(days_in_month * doc_id + 1 + date)
                strin6 = str(book[5]) + str(days_in_month * doc_id + 1 + date)
                strin7 = str(book[6]) + str(days_in_month * doc_id + 1 + date)

                worksheet1[strin1] = str(all_docs[doc_id].get_id())
                worksheet1[strin2] = str(date)



                if doc.get_amount_of_hours(week,day) == 12:
                    worksheet1[strin3] = str('20:00')
                    worksheet1[strin4] = str('9:00')
                    worksheet1[strin5] = str((all_docs[doc_id].get_pereriv(week)[day] % 1 + int(
                        all_docs[doc_id].get_pereriv(week)[day])) * 60)
                    timee = all_docs[doc_id].get_amount_of_hours(week,day) + all_docs[doc_id].get_pereriv(week)[day]

                    worksheet1[strin6] = str(int(timee)) + ':' + str(int((timee % 1) * 60))
                    worksheet1[strin7] = str(doc.get_work_per_day(week,day))

                elif doc.get_amount_of_hours(week,day) == 0:
                    worksheet1[strin3] = str(0)
                    worksheet1[strin4] = str(0)
                else:
                    worksheet1[strin3] = str('8:00')
                    timee = all_docs[doc_id].get_amount_of_hours(week,day) + all_docs[doc_id].get_pereriv(week)[day]
                    worksheet1[strin4] = str(8 + int(timee)) + ':' + str(int((timee % 1) * 60))
                    worksheet1[strin5] = str((all_docs[doc_id].get_pereriv(week)[day] % 1 + int(
                        all_docs[doc_id].get_pereriv(week)[day])) * 60)
                    worksheet1[strin6] = str(int(timee)) + ':' + str(int((timee % 1) * 60))
                    worksheet1[strin7] = str(doc.get_work_per_day(week, day))
                #print(date)
                date = date + 1

        doc_id+=1

    workbook1.save('test21.xlsx')
    df = pd.read_excel('test21.xlsx', header=0)
    os.remove('test21.xlsx')

    conn = sqlite3.connect('db.sqlite3')
    # Запись данных из Excel файла в базу данных
    df.to_sql('accounts_schedule', conn, if_exists='replace', index=True, index_label='id')
    # Завершение работы с базой данных
    conn.commit()
    conn.close()

    zapolnenie_exel(all_docs, amount_for_weeks)

        #print(week)
    #os.remove("test2.xlsx")
    return 0

def make_schetchik_for_all(all_docs):
    new_schetchik = {}
    for doc in all_docs:
        for key in doc.get_abilities():
            if key in new_schetchik.keys():
                new_schetchik[key] += 1
            else:
                new_schetchik[key] = 1
    return new_schetchik

def dop_smena_ravnomerno(all_docs, amount_for_week, activities):
    schetchik = make_schetchik_for_all(all_docs)
    amount_of_hours_per_doc = {}
    for key, value in amount_for_week.items():
        if value > 0:
            amount_of_hours_per_doc[key] = (8 * value) / (activities[key] * schetchik[key])

    return amount_of_hours_per_doc



'''ORDER BY rowid DESC LIMIT 4'''





def get_month_of_week(year, week):
    d = datetime.date(year, 1, 1)
    d = d + datetime.timedelta(weeks=week - 1, days=-d.weekday())
    return d.month


def get_stuff(year, week):
    get_month_of_week(year, week)
    month = get_month_of_week(year, week)

    first_day = datetime.date(year, month, 1)
    first_week_number = first_day.isocalendar()[1]
    return first_week_number


def get_amount_for_weeks(schetchik, activities_UE):
    conn = sqlite3.connect('db.sqlite3')  # Укажите название вашей базы данных
    cursor = conn.cursor()
    my_table = 'accounts_excelmodel'
    year = cursor.execute('SELECT * FROM accounts_excelmodel ORDER BY "index" DESC LIMIT 1 OFFSET 1;')
    year = year.fetchone()[1]
    week = cursor.execute('SELECT * FROM accounts_excelmodel ORDER BY "index" DESC LIMIT 1 OFFSET 1;')
    week = week.fetchone()[2]
    week_to_start_from = get_stuff(year,week)
    #print(week_to_start_from)
    cursor.execute(f'SELECT * FROM accounts_excelmodel ORDER BY "index" DESC LIMIT {week_to_start_from}')
    last_four_rows = cursor.fetchall()

    # Заполнение массива данными
    data_array = []
    for i in range(0, len(last_four_rows)):
        data_array.append(last_four_rows[-1-i])

    amount_of_weeks = []
    a = 0
    for row in data_array:
        i = 3

        amount_of_weeks.append({})
        for key in schetchik.keys():
            if row[i] % 7 < 3.5:
                amount_of_weeks[a][key] = int(row[i] / 7 * activities_UE[key])
            else:
                amount_of_weeks[a][key] = int((int(row[i] / 7) + 1) * activities_UE[key])

            i+=1
        a+=1

    return amount_of_weeks


#181   300   451

def main():
    all_docs = get_doctors_after_correction.create_doctors()
    schetchik = {'Денситометрия': 0, 'КТ': 0, 'КТ1': 0, 'КТ2': 0, 'ММГ': 0, 'МРТ': 0, 'МРТ1': 0, 'МРТ2': 0, 'РГ': 0, 'ФЛГ': 0}
    activities_max_150 = {'Денситометрия': 210, 'КТ': 39, 'КТ1': 24, 'КТ2': 17, 'ММГ': 123, 'МРТ': 30, 'МРТ1': 23,
                          'МРТ2': 15, 'РГ': 123, 'ФЛГ': 451}  # Словарь со всеми исследованиями и их значениями

    activities = {'Денситометрия': 140, 'КТ': 26, 'КТ1': 16, 'КТ2': 11, 'ММГ': 82, 'МРТ': 20, 'МРТ1': 15, 'МРТ2': 10,
                  'РГ': 82, 'ФЛГ': 300}  # Словарь со всеми исследованиями и их значениями
    activities_UE = {'Денситометрия': 2.4, 'КТ': 11.6, 'КТ1': 18.8, 'КТ2': 26.6, 'ММГ': 3.7, 'МРТ': 15.1, 'МРТ1': 19.7, 'МРТ2': 30.1,
                  'РГ': 3.7, 'ФЛГ': 1}  # Словарь со всеми исследованиями и их значениями in UA

    for key in activities_max_150:
        activities[key] = int(activities_max_150[key] / 1.5)
    #print(activities)
    schetchik = make_schetchik(all_docs,schetchik)
    #print(schetchik)

    amount_for_week = get_amount_for_weeks(schetchik, activities_UE)       #2022 15 данные для тестов
    #print(amount_for_week)
    print(make_schedule_for_month(all_docs, activities,activities_UE, amount_for_week, schetchik))


main()
