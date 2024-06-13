import os
import pandas as pd
from statsmodels.tsa.arima.model import ARIMA
import openpyxl
import shutil
import sqlite3

source_file = 'Количество исследований по неделям (для реализации) — копия (2).xlsx'
destination_file = 'Количество исследований с предсказанием.xlsx'
shutil.copy2(source_file, destination_file)


# Загрузка данных из Excel-файла
df = pd.read_excel('Количество исследований по неделям (для реализации) — копия (2).xlsx')
# Удаление строк, в которых значение во второй ячейке равно 1
df = df[df.iloc[:, 1] != 1]
# Сохранение измененного файла
df.to_excel('Количество исследований по неделям без 1 .xlsx', index=False)


t = 0
array_ov = []
for iiii in range(0, 10):
    # Чтение Excel файла
    df = pd.read_excel('Количество исследований по неделям без 1 .xlsx')
    # Выбор третьего столбца без первых двух строк
    column_to_copy = df.iloc[2:, iiii+2]
    # Запись данных в txt файл
    with open(str(iiii+3) + '_Количество исследований по неделям без 1 .txt', 'w') as file:
        for item in column_to_copy:
            file.write(str(item) + '\n')


    # Загружаем данные из файла
    data = pd.read_csv(str(iiii+3) + '_Количество исследований по неделям без 1 .txt', header=None, names=['value'])
    # Обучаем модель ARIMA
    model = ARIMA(data, order=(5,1,1))
    model_fit = model.fit()
    # Предсказываем значения временного ряда
    predictions = model_fit.forecast(steps=5)
    # Выводим предсказанные значения
    arr = []
    # Запись данных в txt файл
    with open(str(iiii+3) + '_Предсказание.txt', 'w') as file:
        for item in predictions:
            file.write(str(item) + '\n')
            arr.append(int(item))
            array_ov.append(int(item))


    file_path = 'Количество исследований с предсказанием.xlsx'
    data = pd.read_excel(file_path)
    num_rows = data.shape[0]
    if iiii == 0:
        num_rows1 = num_rows
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    for i in range(0, len(arr)):
        sheet.cell(row=num_rows1 + 2 + i, column=iiii+3, value=arr[i])
    wb.save(file_path)


rows = []
df = pd.read_excel('Количество исследований с предсказанием.xlsx')
# Поиск строк, в которых значение во втором столбце равно 1
rows_with_1 = df[df.iloc[:, 1] == 1].index
for row in rows_with_1:
    rows.append(row)


# Открываем Excel файл
wb = openpyxl.load_workbook('Количество исследований с предсказанием.xlsx')
sheet = wb.active
# Находим последнюю заполненную строку
last_row = sheet.max_row
for i in range(last_row, 0, -1):
    if sheet.cell(row=i, column=2).value is not None:
        last_value = sheet.cell(row=i, column=2).value
        break


file_path = 'Количество исследований с предсказанием.xlsx'
data = pd.read_excel(file_path)
num_rows = data.shape[0]
wb = openpyxl.load_workbook(file_path)
sheet = wb.active
last_value += 1
for i in range(0, len(arr)):
    if last_value == 53:
        last_value = 1
        sheet.cell(row=num_rows + 2 + i - 5, column=2, value=1)
    if last_value == 1:
        for j in range(0, 10):
            sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=df.iloc[rows[len(rows)-1], j+2])
    if last_value == 2:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-350)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-550)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-50)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-100)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-3000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-500)
            if j == 6:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-100)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-3)
            if j == 8:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-10000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i-5, column=j+3, value=array_ov[i+j*5]-5000)
    if last_value == 8:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 450)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 1550)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 250)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 100)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 6000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 400)
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 300)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 1)
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 15000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 9000)
    if last_value == 18:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 150)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 550)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 110)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 100)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 2000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 100)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 10000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 5000)
    if last_value == 19:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 550)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 950)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 210)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 300)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 5000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 400)
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 300)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 15000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 6000)
    if last_value == 24:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 150)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 350)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 110)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 6000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
    if last_value == 27:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 350)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 650)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 110)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 6000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
    if last_value == 28:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 350)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 650)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 110)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 6000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
    if last_value == 29:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 350)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 650)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 110)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 6000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
    if last_value == 30:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 350)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 650)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 110)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 6000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
    if last_value == 38:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 350)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] + 450)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] + 110)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] + 100)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] + 5000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5])
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 12000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
    if last_value == 52:
        for j in range(0, 10):
            if j == 0:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 350)
            if j == 1:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 650)
            if j == 2:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 3:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 200)
            if j == 4:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
            if j == 5:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 300)
            if j == 6:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 100)
            if j == 7:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 4)
            if j == 8:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 8000)
            if j == 9:
                sheet.cell(row=num_rows + 2 + i - 5, column=j + 3, value=array_ov[i + j * 5] - 3000)
    else:
        sheet.cell(row=num_rows + 2 + i-5, column=2, value=last_value)
    last_value += 1
wb.save(file_path)


os.remove("3_Количество исследований по неделям без 1 .txt")
os.remove("4_Количество исследований по неделям без 1 .txt")
os.remove("5_Количество исследований по неделям без 1 .txt")
os.remove("6_Количество исследований по неделям без 1 .txt")
os.remove("7_Количество исследований по неделям без 1 .txt")
os.remove("8_Количество исследований по неделям без 1 .txt")
os.remove("9_Количество исследований по неделям без 1 .txt")
os.remove("10_Количество исследований по неделям без 1 .txt")
os.remove("11_Количество исследований по неделям без 1 .txt")
os.remove("12_Количество исследований по неделям без 1 .txt")
os.remove("3_Предсказание.txt")
os.remove("4_Предсказание.txt")
os.remove("5_Предсказание.txt")
os.remove("6_Предсказание.txt")
os.remove("7_Предсказание.txt")
os.remove("8_Предсказание.txt")
os.remove("9_Предсказание.txt")
os.remove("10_Предсказание.txt")
os.remove("11_Предсказание.txt")
os.remove("12_Предсказание.txt")


# Чтение данных из Excel файла
df = pd.read_excel('Количество исследований с предсказанием.xlsx')
# Подключение к базе данных SQLite
conn = sqlite3.connect('database.db')
cur = conn.cursor()
# Создание таблицы в базе данных
cur.execute('''CREATE TABLE IF NOT EXISTS data (
                year INTEGER,
                week_number INTEGER,
                densitometer REAL,
                ct REAL,
                ct_with_cu1 REAL,
                ct_with_cu2 REAL,
                mmg REAL,
                mrt REAL,
                mrt_with_cu1 REAL,
                mrt_with_cu2 REAL,
                rg REAL,
                fluorography REAL
                )''')
# Запись данных из Excel файла в базу данных
df.to_sql('data', conn, if_exists='replace', index=False)
# Завершение работы с базой данных
conn.commit()
conn.close()